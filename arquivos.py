
from openpyxl import load_workbook
from itertools import groupby
import re
from pathlib import Path
from config import MES, PASTA_BOLETOS, PASTA_CONDOMINIO, PASTA_REPASSES, ARQ_EXCEL, TXT_ENDERECOS, TXT_NOMES, TXT_EMAILS
from modelos.devedor import Devedores

def ler_txt(path, lista):
    with open(path, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if linha:
                partes = linha.split(". ", 1)
                if len(partes) == 2:
                    lista.append(partes[1])
                else:
                    lista.append(linha)
            else:
                lista.append("")

def extrair_numero(nome_arquivo):
    match = re.search(r'\d+', Path(nome_arquivo).stem)
    return int(match.group()) if match else float('inf')

def processar_pasta(pasta_path):
    arquivos = sorted(pasta_path.glob('*.pdf'), key=extrair_numero)
    grupos = []
    for key, group in groupby(arquivos, key=extrair_numero):
        grupo = list(group)
        grupos.append([str(p) for p in grupo] if len(grupo) > 1 else str(grupo[0]))
    return grupos

def extrair_dados_linha(planilha, x, y):
    dados = {
        'aluguel': None, 'iptu': None, 'cond': None, 'valor': None,
        'cotas_extras': [], 'taxa': None, 'tarifa': None, 'repasse': None
    }
    index = 0

    while planilha.cell(row=x, column=y).value != "Receita":
        celula = planilha.cell(row=x, column=y)
        valor = planilha.cell(row=x, column=y+1).value
        if celula.value == "Aluguel":
            dados['aluguel'] = valor
        elif celula.value == "Iptu cota":
            dados['iptu'] = valor
        elif celula.value == "Taxa de Condomínio":
            dados['cond'] = valor
        elif celula.value == "Valor Boleto":
            dados['valor'] = valor
        elif celula.value is None or "BLT" in celula.value:
            index = 1
        elif celula.value is not None and index == 0 and (float(valor) > 0):
            dados['cotas_extras'].append({celula.value: valor})
        x += 1

    dados['taxa'] = planilha.cell(row=x+2, column=y+1).value
    dados['tarifa'] = planilha.cell(row=x+2, column=y+3).value
    dados['repasse'] = planilha.cell(row=x+3, column=y+1).value
    return dados

def criar_devedor(endereco, nome, email, dados, pdf, cond, repasse):
    return Devedores(
        endereco, nome, email, dados['valor'], dados['aluguel'], dados['iptu'], dados['cond'],
        dados['cotas_extras'], dados['repasse'], dados['taxa'], dados['tarifa'],
        pdf, None if dados['cond'] is not None else cond, repasse
    )

def ler_planilha(n):
    wb = load_workbook(ARQ_EXCEL, data_only=True)
    planilha = wb[MES]
    devedores = []
    imovel = 0
    x = 5
    pdfs = processar_pasta(PASTA_BOLETOS)
    conds = processar_pasta(PASTA_CONDOMINIO)
    repasses = processar_pasta(PASTA_REPASSES)

    enderecos = []
    nomes = []
    emails = []
    ler_txt(TXT_ENDERECOS, enderecos)
    ler_txt(TXT_NOMES, nomes)
    ler_txt(TXT_EMAILS, emails)

    if not (len(enderecos) == len(nomes) == len(emails)):
        raise ValueError("As listas de endereços, nomes e emails devem ter o mesmo tamanho.")

    while imovel < n:
        for y in range(2, 17, 5):
            if imovel >= n:
                break
            dados = extrair_dados_linha(planilha, x, y)
            devedor = criar_devedor(
                enderecos[imovel], nomes[imovel], emails[imovel], dados,
                pdfs[imovel] if imovel < len(pdfs) else None,
                conds[imovel] if imovel < len(conds) else None,
                repasses[imovel] if imovel < len(repasses) else None
            )
            devedores.append(devedor)
            imovel += 1
        x += 17
    wb.close()
    return devedores
